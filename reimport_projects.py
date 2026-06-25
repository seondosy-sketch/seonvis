import openpyxl, json, sys, re, urllib.request, urllib.parse
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('2026_project list.xlsx')
ws = wb.active

def fmt_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    return None  # non-date strings (e.g. '서면') → null

def txt(v):
    if v is None: return ''
    s = str(v).strip()
    return '' if s in ('-', 'ㅡ', 'None') else s

rows = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    if not row[1]:
        continue
    rows.append({
        'project_number': txt(row[1]),   # col2 관리번호
        'type':           txt(row[2]),   # col3 발주방식
        'client':         txt(row[3]),   # col4 발주처
        'name':           txt(row[4]),   # col5 용역명
        'fee':            (float(row[5]) if isinstance(row[5], (int, float)) else None),  # col6
        'tp_score':       txt(row[6]),   # col7 T/P
        'duration_days':  txt(row[7]),   # col8 배점
        'submit_date':    fmt_date(row[8]),   # col9
        'interview_date': fmt_date(row[9]),   # col10 평가일
        'result_score':   txt(row[10]),  # col11 평가결과
        'evaluation':     txt(row[11]),  # col12 최종낙찰사
        'participants':   txt(row[12]),  # col13 참가업체
        'note':           txt(row[13]),  # col14 비고
        'director':       txt(row[14]),  # col15 단장
        'staff_arch':     txt(row[16]),  # col17 건축
        'staff_safety':   txt(row[17]),  # col18 안전
        'staff_civil':    txt(row[18]),  # col19 토목
        'staff_mech':     txt(row[19]),  # col20 기계
    })

env = open('.env.local').read()
SUPABASE_URL = re.search(r'NEXT_PUBLIC_SUPABASE_URL=(.+)', env).group(1).strip()
key = re.search(r'SUPABASE_SERVICE_ROLE_KEY=(.+)', env).group(1).strip()

def patch(pnum, payload):
    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/projects?project_number=eq.{urllib.parse.quote(pnum)}",
        data=data, method='PATCH',
        headers={'apikey': key, 'Authorization': f'Bearer {key}',
                 'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
    )
    with urllib.request.urlopen(req): pass

updated = 0
for r in rows:
    pnum = r.pop('project_number')
    patch(pnum, r)
    updated += 1
    print(f'  [{updated}] {pnum} updated')

print(f'\nDone: {updated} rows updated')
