import openpyxl, json, sys, re, urllib.request, urllib.parse
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('2026_project list.xlsx')
ws = wb.active
rows = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    if not row[1]:
        continue
    rows.append({
        'project_number': str(row[1]).strip(),
        'arch': str(row[16]).strip() if row[16] else '',
        'civil': str(row[17]).strip() if row[17] else '',
        'mech': str(row[18]).strip() if row[18] else '',
        'safety': str(row[19]).strip() if row[19] else '',
    })

env = open('.env.local').read()
SUPABASE_URL = re.search(r'NEXT_PUBLIC_SUPABASE_URL=(.+)', env).group(1).strip()
key = re.search(r'SUPABASE_SERVICE_ROLE_KEY=(.+)', env).group(1).strip()

updated = 0
for r in rows:
    pnum = r['project_number']
    payload = json.dumps({'staff_arch': r['arch'], 'staff_civil': r['civil'], 'staff_mech': r['mech'], 'staff_safety': r['safety']}).encode('utf-8')
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/projects?project_number=eq.{urllib.parse.quote(pnum)}",
        data=payload, method='PATCH',
        headers={'apikey': key, 'Authorization': f'Bearer {key}', 'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
    )
    with urllib.request.urlopen(req) as res:
        updated += 1

print(f'Updated {updated} rows')
